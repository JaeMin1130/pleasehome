#!/usr/bin/env python3
import os
import sys
import argparse
import re
import shutil
import tempfile


def convert_excel_to_md(excel_path):
    """엑셀 파일을 읽어서 각 시트별 데이터를 마크다운 표 문자열로 변환"""
    import openpyxl
    
    print(f"엑셀 변환 시작: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"엑셀 로드 실패: {e}", file=sys.stderr)
        return f"\n\n[오류: 엑셀 파일을 열 수 없습니다. ({os.path.basename(excel_path)})]\n\n"
        
    md_parts = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # 시트 타이틀 추가
        md_parts.append(f"### [주택목록 시트: {sheet_name}]")
        
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            md_parts.append("*빈 시트입니다.*")
            continue
            
        # 첫 행을 헤더로 설정
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        headers = [h.replace("\n", " ").strip() for h in headers]
        headers = [h.replace("|", "\\|") for h in headers]
        
        # 마크다운 표 생성
        md_table = []
        md_table.append("| " + " | ".join(headers) + " |")
        md_table.append("| " + " | ".join(["---"] * len(headers)) + " |")
        
        # 데이터 행
        has_data = False
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            has_data = True
            row_values = [str(cell).replace("\n", " ").replace("|", "\\|").strip() if cell is not None else "" for cell in row]
            if len(row_values) < len(headers):
                row_values.extend([""] * (len(headers) - len(row_values)))
            else:
                row_values = row_values[:len(headers)]
            md_table.append("| " + " | ".join(row_values) + " |")
            
        if not has_data:
            md_parts.append("*데이터가 존재하지 않습니다.*")
        else:
            md_parts.append("\n".join(md_table))
            
    return "\n\n" + "\n\n---\n\n".join(md_parts) + "\n\n"

def convert_single_pdf(pdf_path, output_dir, folder_name):
    # opendataloader-pdf 로드
    import opendataloader_pdf
    
    # 출력 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    print(f"변환 시작: {pdf_path}")
    print(f"변환 출력지: {output_dir}")
    
    pdf_filename = os.path.basename(pdf_path)
    pdf_basename = os.path.splitext(pdf_filename)[0]
    
    # 변환 실행
    opendataloader_pdf.convert(
        input_path=[pdf_path],
        output_dir=output_dir,
        format="markdown"
    )
    
    # 표준화 후처리 경로 지정
    generated_md = os.path.join(output_dir, f"{pdf_basename}.md")
    standard_md = os.path.join(output_dir, "document.md")
    generated_images = os.path.join(output_dir, f"{pdf_basename}_images")
    standard_images = os.path.join(output_dir, "images")

    # 가. 마크다운 파일명 변경
    if os.path.exists(generated_md):
        if os.path.exists(standard_md):
            os.remove(standard_md)
        os.rename(generated_md, standard_md)
        print(f"-> 마크다운 파일 표준화 완료: {standard_md}")
    else:
        md_files = [f for f in os.listdir(output_dir) if f.endswith(".md") and f != "document.md"]
        if md_files:
            shutil.move(os.path.join(output_dir, md_files[0]), standard_md)
            pdf_basename = os.path.splitext(md_files[0])[0]
            generated_images = os.path.join(output_dir, f"{pdf_basename}_images")

    # 나. 이미지 폴더명 변경 및 마크다운 내 이미지 링크 치환
    if os.path.exists(generated_images):
        if os.path.exists(standard_images):
            shutil.rmtree(standard_images)
        os.rename(generated_images, standard_images)
        print(f"-> 이미지 디렉토리 표준화 완료: {standard_images}")

        if os.path.exists(standard_md):
            with open(standard_md, "r", encoding="utf-8") as f:
                content = f.read()
            
            old_rel_path = f"{pdf_basename}_images/"
            new_rel_path = "images/"
            
            if old_rel_path in content:
                content = content.replace(old_rel_path, new_rel_path)
                with open(standard_md, "w", encoding="utf-8") as f:
                    f.write(content)
                print("-> 마크다운 내 이미지 참조 링크 치환 완료.")

def run_auto_scan():
    import json
    base_pdf_dir = "/home/iru/app/pleasehome/db-pipeline/docs/pdf"
    base_md_dir = "/home/iru/app/pleasehome/db-pipeline/docs/md"
    
    # {기관명}_{PAN_ID}_{PAN_DT} 패턴 하위 폴더 스캔
    import re
    notice_folders = []
    for d in os.listdir(base_pdf_dir):
        dir_path = os.path.join(base_pdf_dir, d)
        if os.path.isdir(dir_path):
            # 영문/한글기관명_아이디_8자리날짜 패턴 체크
            if re.match(r'^[A-Za-z가-힣]+_.+_\d{8}$', d):
                notice_folders.append(d)
    
    if not notice_folders:
        print("변환할 공고 폴더(예: LH_{PAN_ID}_{PAN_DT})가 docs/pdf/ 에 없습니다.")
        return
    
    print(f"{len(notice_folders)}개 공고 폴더를 감지했습니다.")
    
    for folder_name in sorted(notice_folders):
        folder_path = os.path.join(base_pdf_dir, folder_name)
        output_dir = os.path.join(base_md_dir, folder_name)
        
        # 이미 변환된 폴더 스킵 (document.md 존재 시)
        if os.path.exists(os.path.join(output_dir, "document.md")):
            print(f"\n[{folder_name}] 이미 변환된 폴더입니다. 건너뜁니다.")
            continue
        
        print(f"\n{'='*50}")
        print(f"공고 폴더 변환 중: {folder_name}")
        
        # download_meta.json 로드
        meta_path = os.path.join(folder_path, "download_meta.json")
        download_meta = {}
        if os.path.exists(meta_path):
            try:
                with open(meta_path, "r", encoding="utf-8") as f:
                    download_meta = json.load(f)
            except Exception as e:
                print(f"[경고] download_meta.json 로드 실패: {e}")
        
        # 변환 대상 파일 수집 (PDF + 엑셀)
        target_files = [
            f for f in os.listdir(folder_path)
            if f.lower().endswith((".pdf", ".xlsx", ".xls"))
        ]
        
        if not target_files:
            print(f"[경고] {folder_name} 폴더에 변환할 파일이 없습니다.")
            continue
        
        print(f"대상 파일 {len(target_files)}개:")
        for f in target_files:
            print(f" - {f}")
        
        os.makedirs(output_dir, exist_ok=True)
        
        # 정렬: 공고문 PDF 우선, 일반 PDF 중간, 엑셀 마지막
        sorted_files = sorted(
            target_files,
            key=lambda x: (
                0 if "공고문" in x or "모집공고" in x else (
                    2 if x.lower().endswith((".xlsx", ".xls")) else 1
                ),
                x
            )
        )
        
        merged_md_parts = []
        final_images_dir = os.path.join(output_dir, "images")
        
        for idx, file_name in enumerate(sorted_files):
            file_path = os.path.join(folder_path, file_name)
            
            # 엑셀 파일 변환
            if file_name.lower().endswith((".xlsx", ".xls")):
                try:
                    excel_md = convert_excel_to_md(file_path)
                    merged_md_parts.append(excel_md)
                    print(f"-> 엑셀 변환 완료: {file_name}")
                except Exception as e:
                    print(f"-> 엑셀 {file_name} 변환 실패: {e}", file=sys.stderr)
                continue
            
            # PDF 변환
            temp_out = os.path.join(output_dir, f"_temp_{idx}")
            try:
                convert_single_pdf(file_path, temp_out, folder_name)
                
                temp_md_path = os.path.join(temp_out, "document.md")
                if os.path.exists(temp_md_path):
                    with open(temp_md_path, "r", encoding="utf-8") as md_f:
                        part_content = md_f.read()
                    
                    temp_images_dir = os.path.join(temp_out, "images")
                    if os.path.exists(temp_images_dir):
                        os.makedirs(final_images_dir, exist_ok=True)
                        for img_name in os.listdir(temp_images_dir):
                            src_img = os.path.join(temp_images_dir, img_name)
                            new_img_name = f"part_{idx}_{img_name}"
                            dst_img = os.path.join(final_images_dir, new_img_name)
                            shutil.move(src_img, dst_img)
                            part_content = part_content.replace(f"images/{img_name}", f"images/{new_img_name}")
                    
                    merged_md_parts.append(part_content)
                else:
                    print(f"-> 경고: {file_name} 변환 결과 마크다운 없음", file=sys.stderr)
            except Exception as e:
                print(f"-> {file_name} 변환 실패: {e}", file=sys.stderr)
            finally:
                if os.path.exists(temp_out):
                    shutil.rmtree(temp_out)
        
        if not merged_md_parts:
            print(f"-> 오류: 변환된 마크다운이 없습니다.", file=sys.stderr)
            continue
        
        # 최종 document.md 병합 저장
        final_md_path = os.path.join(output_dir, "document.md")
        final_content = "\n\n---\n\n".join(merged_md_parts)
        with open(final_md_path, "w", encoding="utf-8") as f:
            f.write(final_content)
        print(f"-> 최종 변환 성공: docs/md/{folder_name}/document.md")
        
        # api_meta.json 생성 (하이브리드 포맷)
        if download_meta:
            api_meta_path = os.path.join(output_dir, "api_meta.json")
            try:
                # 전세임대 여부에 따른 has_complexes 동적 판별
                has_complexes = True
                ais_tp_cd = download_meta.get("AIS_TP_CD", "")
                ais_tp_cd_nm = download_meta.get("AIS_TP_CD_NM", "")
                if ais_tp_cd == "062" or "전세임대" in ais_tp_cd_nm:
                    has_complexes = False
                
                # 청약 유형 정규화
                category = ais_tp_cd_nm or "공공임대"
                if "행복주택" in category:
                    category = "행복주택"
                elif "매입임대" in category:
                    category = "매입임대"
                elif "든든전세" in category:
                    category = "든든전세"
                elif "전세임대" in category:
                    category = "전세임대"
                
                api_meta_content = {
                    "pan_id": download_meta.get("PAN_ID", ""),
                    "title": download_meta.get("PAN_NM", ""),
                    "api_metadata": download_meta,
                    "inferred_standards": {
                        "subscription_type": category,
                        "institution": download_meta.get("_institution", "LH"),
                        "has_complexes": has_complexes,
                        "is_distributed": False
                    }
                }
                
                with open(api_meta_path, "w", encoding="utf-8") as af:
                    json.dump(api_meta_content, af, ensure_ascii=False, indent=2)
                print(f"-> api_meta.json 하이브리드 포맷 생성 성공: docs/md/{folder_name}/api_meta.json")
            except Exception as ae:
                print(f"[경고] api_meta.json 생성 실패: {ae}", file=sys.stderr)

def main():
    parser = argparse.ArgumentParser(
        description="PDF 파일을 표준 마크다운(document.md & images/)으로 변환하고 정리하는 통합 스크립트입니다."
    )
    parser.add_argument(
        "target",
        nargs="?",
        default=None,
        help="변환할 공고 폴더명(예: LH_2015122300020216_20260625) 또는 PDF 직접 경로 (생략 시 자동 미분류 파일 스캔)"
    )
    
    args = parser.parse_args()
    
    # 인자가 생략되었으면 자동 스캔 실행
    if args.target is None:
        run_auto_scan()
        sys.exit(0)
        
    # 인자가 주어졌으면 단일 파일 변환 모드로 가동
    base_pdf_dir = "/home/iru/app/pleasehome/db-pipeline/docs/pdf"
    base_md_dir = "/home/iru/app/pleasehome/db-pipeline/docs/md"
    
    target = args.target
    if not os.path.exists(target) and not os.path.isabs(target):
        folder_pdf_dir = os.path.join(base_pdf_dir, target)
        if os.path.isdir(folder_pdf_dir):
            pdfs = [f for f in os.listdir(folder_pdf_dir) if f.lower().endswith(".pdf")]
            if pdfs:
                pdf_path = os.path.join(folder_pdf_dir, pdfs[0])
            else:
                pdf_path = os.path.join(folder_pdf_dir, "origin.pdf")  # fallback
        else:
            pdf_path = os.path.join(base_pdf_dir, target)
        output_dir = os.path.join(base_md_dir, target)
        folder_name = target
    else:
        pdf_path = os.path.abspath(target)
        folder_name = os.path.basename(os.path.dirname(pdf_path))
        output_dir = os.path.join(base_md_dir, folder_name)

    if not os.path.isfile(pdf_path):
        print(f"오류: PDF 파일을 찾을 수 없습니다: {pdf_path}", file=sys.stderr)
        sys.exit(1)
        
    try:
        convert_single_pdf(pdf_path, output_dir, folder_name)
        print("모든 변환 및 표준화 후처리가 성공적으로 완료되었습니다!")
    except Exception as e:
        print(f"변환 중 오류 발생: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()

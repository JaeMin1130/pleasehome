#!/usr/bin/env python3
import os
import sys
import argparse
import re
import shutil
import tempfile


def convert_excel_to_md(excel_path):
    """엑셀 파일을 읽어서 각 시트별 데이터를 마크다운 표 문자열로 변환"""
    import openpyxl
    
    print(f"엑셀 변환 시작: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"엑셀 로드 실패: {e}", file=sys.stderr)
        return f"\n\n[오류: 엑셀 파일을 열 수 없습니다. ({os.path.basename(excel_path)})]\n\n"
        
    md_parts = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # 시트 타이틀 추가
        md_parts.append(f"### [주택목록 시트: {sheet_name}]")
        
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            md_parts.append("*빈 시트입니다.*")
            continue
            
        # 첫 행을 헤더로 설정
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        headers = [h.replace("\n", " ").strip() for h in headers]
        headers = [h.replace("|", "\\|") for h in headers]
        
        # 마크다운 표 생성
        md_table = []
        md_table.append("| " + " | ".join(headers) + " |")
        md_table.append("| " + " | ".join(["---"] * len(headers)) + " |")
        
        # 데이터 행
        has_data = False
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            has_data = True
            row_values = [str(cell).replace("\n", " ").replace("|", "\\|").strip() if cell is not None else "" for cell in row]
            if len(row_values) < len(headers):
                row_values.extend([""] * (len(headers) - len(row_values)))
            else:
                row_values = row_values[:len(headers)]
            md_table.append("| " + " | ".join(row_values) + " |")
            
        if not has_data:
            md_parts.append("*데이터가 존재하지 않습니다.*")
        else:
            md_parts.append("\n".join(md_table))
            
    return "\n\n" + "\n\n---\n\n".join(md_parts) + "\n\n"

def parse_metadata(pdf_path):
    filename = os.path.basename(pdf_path)
    
    # 1. 파일명에서 연도 추출 (4자리 숫자)
    year = "2026"  # 기본값
    year_match = re.search(r"(20\d{2})", filename)
    if year_match:
        year = year_match.group(1)
        
    # 2. 파일명에서 차수/구분 추출
    order = "1차"  # 기본값
    if "상반기" in filename:
        order = "상반기"
    elif "하반기" in filename:
        order = "하반기"
    elif "특화형" in filename:
        order = "특화형"
    elif "공임50년" in filename or "50년공공임대" in filename:
        order = "50년"
    else:
        order_match = re.search(r"제?(\d+)차", filename)
        if order_match:
            order = f"{order_match.group(1)}차"
        elif "신혼신생아" in filename:
            order = "신혼신생아1"
            
    # 특수 처리: 자립준비청년 키워드 결합
    if "자립준비청년" in filename:
        order = f"{order}_자립준비청년"
        
    # 공고 개별 분리 보장을 위해 세부 구분 추가
    if "신혼" in filename or "신생아" in filename:
        if "Ⅰ" in filename or "1" in filename:
            order = f"{order}_신혼신생아1"
        elif "Ⅱ" in filename or "2" in filename:
            order = f"{order}_신혼신생아2"
        else:
            order = f"{order}_신혼신생아"
    elif "청년" in filename:
        order = f"{order}_청년"
    elif "기숙사" in filename:
        order = f"{order}_기숙사형"
            
    # 3. 파일명에서 청약유형 추출
    category = None
    if "행복주택" in filename:
        category = "행복주택"
    elif "장기전세주택2" in filename or "장기전세2" in filename or "미리내집" in filename:
        category = "장기전세2"
    elif "장기전세" in filename:
        category = "장기전세"
    elif "든든전세" in filename or "든든주택" in filename:
        category = "든든전세"
    elif "국민임대" in filename:
        category = "국민임대"
    elif "영구임대" in filename:
        category = "영구임대"
    elif "통합공공임대" in filename:
        category = "통합공공임대"
    elif "매입임대" in filename:
        category = "매입임대"
    elif "전세임대" in filename:
        category = "전세임대"
    elif "청년안심" in filename:
        category = "청년안심"
    elif "장기안심" in filename:
        category = "장기안심"
    elif "희망하우징" in filename:
        category = "희망하우징"
    elif "공공임대" in filename:
        category = "공공임대"

    # 4. 기관명 1차 추출 (LH, SH, GH, iH, HUG, 민간) - 대소문자 무관 및 명칭 보완
    institution = None
    filename_lower = filename.lower()
    if "민간" in filename_lower:
        institution = "민간"
    elif "lh" in filename_lower or "한국토지주택" in filename_lower:
        institution = "LH"
    elif "sh" in filename_lower or "서울주택도시" in filename_lower:
        institution = "SH"
    elif "gh" in filename_lower or "경기주택도시" in filename_lower or "경기도시" in filename_lower:
        institution = "GH"
    elif "ih" in filename_lower or "인천도시" in filename_lower:
        institution = "iH"
    elif "hug" in filename_lower or "주택도시보증" in filename_lower:
        institution = "HUG"
    else:
        # 파일명에서 지자체명/조합명 등 감지 시도 (백업)
        agency_match = re.search(r"\(?[a-zA-Z0-9_]*\)?([가-힣\w]+(?:도시공사|주택공사|협동조합|시|군|구))", filename)
        if agency_match:
            institution = agency_match.group(1)

    # 5. 본문 텍스트 분석 기반 보정 (기관명이 표준 5대 기관이 아니거나 유형을 감지 못한 경우 강제 구동)
    if not category or institution not in ["LH", "SH", "GH", "iH", "HUG", "민간"]:
        try:
            import opendataloader_pdf
            with tempfile.TemporaryDirectory() as temp_dir:
                opendataloader_pdf.convert(
                    input_path=[pdf_path],
                    output_dir=temp_dir,
                    format="markdown"
                )
                md_files = [f for f in os.listdir(temp_dir) if f.endswith(".md")]
                if md_files:
                    with open(os.path.join(temp_dir, md_files[0]), "r", encoding="utf-8") as f:
                        text = f.read(5000)  # 첫 5000자 분석
                    
                    if not category:
                        if "행복주택" in text:
                            category = "행복주택"
                        elif "장기전세" in text:
                            if "장기전세주택2" in text or "미리내집" in text:
                                category = "장기전세2"
                            else:
                                category = "장기전세"
                        elif "든든전세" in text or "든든주택" in text:
                            category = "든든전세"
                        elif "국민임대" in text:
                            category = "국민임대"
                        elif "영구임대" in text:
                            category = "영구임대"
                        elif "통합공공임대" in text:
                            category = "통합공공임대"
                        elif "매입임대" in text:
                            category = "매입임대"
                        elif "전세임대" in text:
                            category = "전세임대"
                        elif "청년안심" in text:
                            category = "청년안심"
                        elif "장기안심" in text:
                            category = "장기안심"
                    
                    # 기관명 재검증 및 강제 보정 - 대소문자 무관 및 명칭 보완
                    text_lower = text.lower()
                    if "민간" in filename_lower or "민간임대" in text_lower or "민간임대주택" in text_lower:
                        institution = "민간"
                    elif "한국토지주택" in text_lower or "lh" in text_lower or "l.h" in text_lower:
                        institution = "LH"
                    elif "서울주택도시" in text_lower or "sh" in text_lower:
                        institution = "SH"
                    elif "경기주택도시" in text_lower or "gh" in text_lower or "경기도시" in text_lower:
                        institution = "GH"
                    elif "인천도시" in text_lower or "ih" in text_lower:
                        institution = "iH"
                    elif "주택도시보증" in text_lower or "hug" in text_lower:
                        institution = "HUG"
        except Exception as e:
            pass

    # 최종 기본값 세팅
    if not institution or institution not in ["LH", "SH", "GH", "iH", "HUG", "민간"]:
        institution = "LH"
    if not category:
        category = "공공임대"
        
    return year, order, category, institution

def convert_single_pdf(pdf_path, output_dir, folder_name):
    # opendataloader-pdf 로드
    import opendataloader_pdf
    
    # 출력 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    print(f"변환 시작: {pdf_path}")
    print(f"변환 출력지: {output_dir}")
    
    pdf_filename = os.path.basename(pdf_path)
    pdf_basename = os.path.splitext(pdf_filename)[0]
    
    # 변환 실행
    opendataloader_pdf.convert(
        input_path=[pdf_path],
        output_dir=output_dir,
        format="markdown"
    )
    
    # 표준화 후처리 (document.md & images/ 로 리네임)
    generated_md = os.path.join(output_dir, f"{pdf_basename}.md")
    standard_md = os.path.join(output_dir, "document.md")
    generated_images = os.path.join(output_dir, f"{pdf_basename}_images")
    standard_images = os.path.join(output_dir, "images")

    # 가. 마크다운 파일명 변경
    if os.path.exists(generated_md):
        if os.path.exists(standard_md):
            os.remove(standard_md)
        os.rename(generated_md, standard_md)
        print(f"-> 마크다운 파일 표준화 완료: {standard_md}")
    else:
        # 예외 상황: 변환된 다른 이름의 md가 있는지 탐색
        md_files = [f for f in os.listdir(output_dir) if f.endswith(".md") and f != "document.md"]
        if md_files:
            shutil.move(os.path.join(output_dir, md_files[0]), standard_md)
            pdf_basename = os.path.splitext(md_files[0])[0]
            generated_images = os.path.join(output_dir, f"{pdf_basename}_images")

    # 나. 이미지 폴더명 변경 및 마크다운 내 이미지 링크 치환
    if os.path.exists(generated_images):
        if os.path.exists(standard_images):
            shutil.rmtree(standard_images)
        os.rename(generated_images, standard_images)
        print(f"-> 이미지 디렉토리 표준화 완료: {standard_images}")

        # 마크다운 파일 내의 이미지 참조 경로 치환 (예: origin_images/ -> images/)
        if os.path.exists(standard_md):
            with open(standard_md, "r", encoding="utf-8") as f:
                content = f.read()
            
            old_rel_path = f"{pdf_basename}_images/"
            new_rel_path = "images/"
            
            if old_rel_path in content:
                content = content.replace(old_rel_path, new_rel_path)
                with open(standard_md, "w", encoding="utf-8") as f:
                    f.write(content)
                print("-> 마크다운 내 이미지 참조 링크 치환 완료.")

def run_auto_scan():
    base_pdf_dir = "/home/iru/app/pleasehome/backend/docs/pdf"
    base_md_dir = "/home/iru/app/pleasehome/backend/docs/md"
    
    # 1. 미분류 PDF 및 엑셀 파일 목록 가져오기
    raw_files = [
        f for f in os.listdir(base_pdf_dir) 
        if f.lower().endswith((".pdf", ".xlsx", ".xls")) and os.path.isfile(os.path.join(base_pdf_dir, f))
    ]
    
    if not raw_files:
        print("정리할 미분류 PDF 또는 엑셀 파일이 docs/pdf/ 루트에 존재하지 않습니다.")
        return
        
    print(f"미분류 파일 {len(raw_files)}개를 감지했습니다.")
    
    # 가. 메타데이터 파싱 및 그룹화
    from collections import defaultdict
    groups = defaultdict(list)
    for file_name in raw_files:
        file_path = os.path.join(base_pdf_dir, file_name)
        # 엑셀 파일명 파싱도 parse_metadata로 지원
        year, order, category, temp_inst = parse_metadata(file_path)
        groups[(year, order, category)].append({
            "name": file_name,
            "path": file_path,
            "temp_inst": temp_inst
        })
        
    for (year, order, category), files in groups.items():
        print(f"\n==========================================")
        print(f"그룹 처리 중: {year}_{order}_{category} (파일 수: {len(files)})")
        for f in files:
            print(f" - {f['name']} (가판정 기관: {f['temp_inst']})")
            
        # 나. 대표 기관명 결정
        final_inst = None
        anchor_files = [f for f in files if "공고문" in f["name"] or "모집공고" in f["name"]]
        if anchor_files:
            for f in anchor_files:
                if f["temp_inst"] in ["LH", "SH", "GH", "iH", "HUG", "민간"]:
                    final_inst = f["temp_inst"]
                    break
            if not final_inst:
                final_inst = anchor_files[0]["temp_inst"]
        if not final_inst:
            for f in files:
                if f["temp_inst"] in ["LH", "SH", "GH", "iH", "HUG", "민간"]:
                    final_inst = f["temp_inst"]
                    break
        if not final_inst:
            final_inst = files[0]["temp_inst"] or "LH"
            
        # 다. 시퀀스 번호 결정 (동일 조합 존재 시 자동 증가)
        prefix = f"{year}_{order}_{category}_{final_inst}_"
        existing_seqs = []
        for d in os.listdir(base_pdf_dir):
            if os.path.isdir(os.path.join(base_pdf_dir, d)) and d.startswith(prefix):
                try:
                    seq_str = d[len(prefix):]
                    existing_seqs.append(int(seq_str))
                except ValueError:
                    pass
        
        if existing_seqs:
            next_seq = max(existing_seqs) + 1
        else:
            next_seq = 1
            
        std_folder_name = f"{year}_{order}_{category}_{final_inst}_{next_seq:02d}"
        print(f"-> 판정된 표준 공고명: {std_folder_name}")
        
        target_folder = os.path.join(base_pdf_dir, std_folder_name)
        os.makedirs(target_folder, exist_ok=True)
        
        # 라. 파일 이동 (PDF 및 엑셀 모두 타겟 폴더로 이동)
        for f in files:
            target_pdf_path = os.path.join(target_folder, f["name"])
            shutil.move(f["path"], target_pdf_path)
            print(f"-> 표준 폴더로 이동 완료: {target_pdf_path}")
            f["target_pdf_path"] = target_pdf_path
            
        # 마. 변환 및 결과 병합
        output_dir = os.path.join(base_md_dir, std_folder_name)
        os.makedirs(output_dir, exist_ok=True)
        
        # 순서 정렬: 공고문 PDF 우선, 일반 PDF 중간, 엑셀 파일은 마지막에 오도록 배치
        sorted_files = sorted(
            files,
            key=lambda x: (
                0 if "공고문" in x["name"] or "모집공고" in x["name"] else (
                    2 if x["name"].lower().endswith((".xlsx", ".xls")) else 1
                ),
                x["name"]
            )
        )
        
        merged_md_parts = []
        final_images_dir = os.path.join(output_dir, "images")
        
        # 임시 변환 디렉토리들 생성 후 변환
        for idx, f in enumerate(sorted_files):
            # 엑셀 파일의 경우 마크다운 표 변환 처리
            if f["target_pdf_path"].lower().endswith((".xlsx", ".xls")):
                try:
                    excel_md = convert_excel_to_md(f["target_pdf_path"])
                    merged_md_parts.append(excel_md)
                    print(f"-> 엑셀 파일 마크다운 표 변환 완료: {f['name']}")
                except Exception as e:
                    print(f"-> 엑셀 {f['name']} 변환 실패! 에러: {e}", file=sys.stderr)
                continue
                
            temp_out = os.path.join(output_dir, f"_temp_{idx}")
            try:
                # 개별 PDF를 임시 폴더로 변환
                convert_single_pdf(f["target_pdf_path"], temp_out, std_folder_name)
                
                temp_md_path = os.path.join(temp_out, "document.md")
                if os.path.exists(temp_md_path):
                    with open(temp_md_path, "r", encoding="utf-8") as md_f:
                        part_content = md_f.read()
                        
                    # 이미지 정리 및 경로 치환
                    temp_images_dir = os.path.join(temp_out, "images")
                    if os.path.exists(temp_images_dir):
                        os.makedirs(final_images_dir, exist_ok=True)
                        for img_name in os.listdir(temp_images_dir):
                            src_img = os.path.join(temp_images_dir, img_name)
                            # 중복 방지를 위한 이미지명 접두사화
                            new_img_name = f"part_{idx}_{img_name}"
                            dst_img = os.path.join(final_images_dir, new_img_name)
                            shutil.move(src_img, dst_img)
                            
                            # 마크다운 내 이미지 경로 치환
                            part_content = part_content.replace(f"images/{img_name}", f"images/{new_img_name}")
                            
                    merged_md_parts.append(part_content)
                else:
                    print(f"-> 경고: {f['name']}의 마크다운 파일 변환 결과를 찾을 수 없습니다.", file=sys.stderr)
            except Exception as e:
                print(f"-> {f['name']} 변환 실패! 에러: {e}", file=sys.stderr)
            finally:
                # 임시 디렉토리 삭제
                if os.path.exists(temp_out):
                    shutil.rmtree(temp_out)
                    
        # 최종 파일 저장
        if merged_md_parts:
            final_md_path = os.path.join(output_dir, "document.md")
            final_content = "\n\n---\n\n".join(merged_md_parts)
            with open(final_md_path, "w", encoding="utf-8") as final_md_f:
                final_md_f.write(final_content)
            print(f"-> 최종 변환 성공! 병합 경로: docs/md/{std_folder_name}/document.md")
        else:
            print(f"-> 오류: 변환된 결과 마크다운이 존재하지 않습니다.", file=sys.stderr)

def main():
    parser = argparse.ArgumentParser(
        description="PDF 파일을 표준 마크다운(document.md & images/)으로 변환하고 정리하는 통합 스크립트입니다."
    )
    parser.add_argument(
        "target",
        nargs="?",
        default=None,
        help="변환할 공고 폴더명(예: 2026_1차_행복주택) 또는 PDF 직접 경로 (생략 시 자동 미분류 파일 스캔)"
    )
    
    args = parser.parse_args()
    
    # 인자가 생략되었으면 자동 스캔 실행
    if args.target is None:
        run_auto_scan()
        sys.exit(0)
        
    # 인자가 주어졌으면 단일 파일 변환 모드로 가동
    base_pdf_dir = "/home/iru/app/pleasehome/backend/docs/pdf"
    base_md_dir = "/home/iru/app/pleasehome/backend/docs/md"
    
    target = args.target
    if not os.path.exists(target) and not os.path.isabs(target):
        folder_pdf_dir = os.path.join(base_pdf_dir, target)
        if os.path.isdir(folder_pdf_dir):
            pdfs = [f for f in os.listdir(folder_pdf_dir) if f.lower().endswith(".pdf")]
            if pdfs:
                pdf_path = os.path.join(folder_pdf_dir, pdfs[0])
            else:
                pdf_path = os.path.join(folder_pdf_dir, "origin.pdf")  # fallback
        else:
            pdf_path = os.path.join(base_pdf_dir, target)
        output_dir = os.path.join(base_md_dir, target)
        folder_name = target
    else:
        pdf_path = os.path.abspath(target)
        folder_name = os.path.basename(os.path.dirname(pdf_path))
        output_dir = os.path.join(base_md_dir, folder_name)

    if not os.path.isfile(pdf_path):
        print(f"오류: PDF 파일을 찾을 수 없습니다: {pdf_path}", file=sys.stderr)
        sys.exit(1)
        
    try:
        convert_single_pdf(pdf_path, output_dir, folder_name)
        print("모든 변환 및 표준화 후처리가 성공적으로 완료되었습니다!")
    except Exception as e:
        print(f"변환 중 오류 발생: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
